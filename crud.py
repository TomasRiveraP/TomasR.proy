from datetime import datetime
from openpyxl import load_workbook

def leer(ruta:str, extraer:str):
    archivo_excel = load_workbook(ruta)
    Hoja_datos = archivo_excel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:

        if ininstance(i[0].value, int):
            info.setdefault(i[0].value, {'tarea':i[1].value, 'descripcion':i[2].value, 'estado':i[3].value, 'fecha de inicio':i[4].value, 'fecha de finalizacion':[5].value })
    if not(extraer=='todo'):
        info=filtrar(info, extraer)

    for i in info:
        print('******** Tarea *******')
        print('Id:'+ str(i)+'\n'+ 'Titulo: '+str(info[i]['tarea'])+'\n'+'Descripcion: '+str(info[i]['descripcion']) + '\n'+ 'Estado: '+str(info[i]['estado']) +'\n'+ 'Fecha Creacion: ' +str(info[i]['fecha de inicio']) +'\n'+ 'Fecha de finalizacion: ' +str(info[i]['fecha de finalizacion']) )
        print()

    return


def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    return aux

def actualizar(ruta:str, identificador:int,datos_actualizados:dict):
    archivo_excel = load_workbook(ruta)
    Hoja_datos = archivo_excel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=archivo_excel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_Finalizado=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=='titulo' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=titulo).value=datosActualizados[d]
                elif d=='descripcion' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=descripcion).value=datosActualizados[d]
                elif d=='estado' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=estado).value=datosActualizados[d]
                elif d=='fecha inicio' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=fecha_inicio).value=datosActualizados[d]
                elif d=='fecha finalizacion' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=fecha_Finalizado).value=datosActualizados[d]
    archivo_excel.save(ruta)
    if encontro==False:
        print('error: no existeuna tarea con ese Id')
        print()
    return

def agregar(ruta:int, datos:dict):
    archivo_excel = load_workbook(ruta)
    Hoja_datos = archivo_excel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]
    hoja=archivo_excel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_Finalizado=6
    for i in Hoja_datos:
        if not( isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador,column=titulo).value=datos['titulo']
            hoja.cell(row=identificador,column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador,column=estado).value=datos['estado']
            hoja.cell(row=identificador,column=fecha_inicio).value=datos['fecha inicio']
            hoja.cell(row=identificador,column=fecha_Finalizado).value=datos['fecha finalizacion']
            break
    archivo_excel.save(ruta)
    return

def borrar(ruta,identificador):
    archivo_excel = load_workbook(ruta)
    Hoja_datos = archivo_excel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=archivo_excel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_Finalizado=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row.fila, column=1). value=""
            hoja.cell(row.fila, column=titulo). value=""
            hoja.cell(row.fila, column=descripcion). value=""
            hoja.cell(row.fila, column=estado). value=""
            hoja.cell(row.fila, column=fecha_inicio). value=""
            hoja.cell(row.fila, column=fecha_Finalizado). value=""
    archivo_excel.save(ruta)
    if encontro==False:
        print('error: No existe una tarea con ese Id')
        print()
    return

datosActualizados={'titulo':'', 'descripcion':'', 'estado':'', 'fecha inicio':'', 'fecha finalizacion':''}
while True:
    print('Indique la accion que desea realizar')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear nueva tarea: 3')
    print('Borrar: 4')
    accion = input('Escriba la opcion')

    
ruta ="C:\\Users\\user\\OneDrive\\Escritorio\\python\\intro\\dbTask.xlsx"
ruta =r"C:\Users\user\OneDrive\Escritorio\python\intro\dbTask.xlsx"